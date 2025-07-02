import os
import re
import argparse
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string



def process_range(range_str):
    """解析范围字符串并返回处理区域 (支持多字母列名)"""
    pattern = r"([A-Za-z]+)(\d+):([A-Za-z]+)(\d+)"
    match = re.match(pattern, range_str)
    if not match:
        raise ValueError(
            f"范围格式错误: '{range_str}'，请使用类似 'A1:B2' 或 'A2:NL200' 的格式"
        )

    start_col_letter, start_row_str, end_col_letter, end_row_str = match.groups()

    start_col = column_index_from_string(start_col_letter)
    start_row = int(start_row_str)
    end_col = column_index_from_string(end_col_letter)
    end_row = int(end_row_str)

    return (start_row, start_col, end_row, end_col)


def fill_columns(ws, range_coords):
    """填充指定范围内的列数据"""
    start_row, start_col, end_row, end_col = range_coords

    # 处理第一行中的每个单元格
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=start_row, column=col)
        if cell.value is not None and cell.value != "":
            # 用相同的值填充当前列的所有单元格
            for row in range(start_row + 1, end_row + 1):
                target_cell = ws.cell(row=row, column=col)
                target_cell.value = cell.value


def process_excel_file(file_path, range_str, suffix, sheet_name):
    """处理单个Excel文件"""
    try:
        wb = load_workbook(file_path)
    except Exception as e:
        print(f"加载文件 {file_path} 失败: {str(e)}")
        return

    # 检查指定的Sheet是否存在
    if sheet_name not in wb.sheetnames:
        print(f"警告: 工作表 '{sheet_name}' 在文件 {file_path} 中不存在。跳过处理。")
        return

    try:
        # 处理范围
        range_coords = process_range(range_str)
    except ValueError as e:
        print(f"错误: {str(e)}")
        return

    # 只处理指定的工作表
    ws = wb[sheet_name]

    # 验证范围是否在有效区域内
    # if range_coords[2] > ws.max_row or range_coords[3] > ws.max_column:
    #     print(
    #         f"警告: 范围 {range_str} 超出工作表 '{sheet_name}' 的边界 ({ws.max_row}行, {ws.max_column}列)。跳过处理。"
    #     )
    #     return

    try:
        fill_columns(ws, range_coords)
    except Exception as e:
        print(f"处理文件 {file_path} 时出错: {str(e)}")
        return

    # 保存处理后的文件
    base, ext = os.path.splitext(file_path)
    new_file_path = f"{base}{suffix}{ext}"
    try:
        wb.save(new_file_path)
        print(f"已处理并保存: {new_file_path}")
    except Exception as e:
        print(f"保存文件 {file_path} 失败: {str(e)}")


if __name__ == "__main__":
    # process_excel_file(
    #     file_path="src/test/col_stretch/拉伸测试.xlsx",
    #     range_str="A3:NL200",
    #     sheet_name="待处理",
    #     suffix="_processed",
    # )
    parser = argparse.ArgumentParser()

    parser.add_argument("--fp", required=True, help="要处理的Excel文件路径")
    parser.add_argument("--xy", required=True, help="列范围，格式如'A1:B100'")
    parser.add_argument(
        "--st", default="Sheet1", help="要处理的工作表名，默认为'Sheet1'"
    )
    parser.add_argument(
        "--suf", default="_streched", help='处理后文件的后缀，默认为"_streched"'
    )
    args = parser.parse_args()
    process_excel_file(
        file_path=args.fp,
        range_str=args.xy,
        sheet_name=args.st,
        suffix=args.suf,
    )
"""
python D:\my-project\erp-excel-tools\src\scripts\col_stretch.py `
--fp "D:\my-project\erp-excel-tools\src\test\col_stretch\拉伸测试.xlsx" `
--xy "A3:NL10"
"""