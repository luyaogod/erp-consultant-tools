import sqlite3
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet
from typing import List, Tuple, Callable, Optional
from pathlib import Path
from pydantic import BaseModel, field_validator, Field
from pydantic_core.core_schema import FieldValidationInfo


class DatabaseManager:
    """数据库管理"""

    def __init__(self, db_path: str = "excel_processor.db"):
        self.db_path = db_path
        self.conn = None
        self._ensure_database()

    def _ensure_database(self):
        """确保数据库文件存在"""
        # 检查目录是否存在，不存在则创建
        db_dir = os.path.dirname(self.db_path) or "."
        os.makedirs(db_dir, exist_ok=True)

        # 连接数据库（如果文件不存在会自动创建）
        self.conn = sqlite3.connect(self.db_path)
        self._create_table_if_not_exists()

    def _create_table_if_not_exists(self):
        """确保表存在"""
        # 检查表是否存在
        cursor = self.conn.cursor()
        cursor.execute("""
        SELECT count(name) FROM sqlite_master 
        WHERE type='table' AND name='processed_records'
        """)

        # 如果表不存在则创建
        if cursor.fetchone()[0] == 0:
            self.conn.execute("""
            CREATE TABLE processed_records (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                code TEXT NOT NULL,
                serial INTEGER NOT NULL,
                create_time TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE (code, serial)
            )
            """)
            self.conn.commit()

    def begin_transaction(self):
        """开始事务"""
        self.conn.execute("BEGIN TRANSACTION")

    def commit_transaction(self):
        """提交事务"""
        self.conn.commit()

    def rollback_transaction(self):
        """回滚事务"""
        self.conn.rollback()

    def get_max_serial(self, code: str) -> Optional[int]:
        """获取指定编码的最大流水号"""
        cursor = self.conn.execute(
            "SELECT MAX(serial) FROM processed_records WHERE code = ?",
            (code,),
        )
        return cursor.fetchone()[0]

    def insert_records_with_begin_serial(
        self,
        codes: List[str],
        begin_serial: Optional[int] = None,
    ) -> List[Tuple[str, int]]:
        """批量插入记录并返回流水号列表

        Args:
            codes: 编码列表
            begin_serial: 指定的起始流水号，如果不指定则自动递增

        Returns:
            返回包含编码和流水号的元组列表

        Raises:
            ValueError: 如果指定的流水号与已有记录冲突
        """
        results = []

        try:
            self.begin_transaction()

            if begin_serial is not None:
                # 检查所有编码是否都可用指定的起始流水号
                for i, code in enumerate(codes):
                    serial = begin_serial + i

                    # 检查该流水号是否已存在
                    cursor = self.conn.execute(
                        "SELECT 1 FROM processed_records WHERE code = ? AND serial = ?",
                        (code, serial),
                    )
                    if cursor.fetchone() is not None:
                        max_serial = self.get_max_serial(code)
                        raise ValueError(
                            f"编码 '{code}' 的流水号 {serial} 已存在，当前最大流水号为 {max_serial}"
                        )

            # 插入所有记录
            for i, code in enumerate(codes):
                if begin_serial is not None:
                    serial = begin_serial + i
                else:
                    # 自动递增模式
                    max_serial = self.get_max_serial(code)
                    serial = 1 if max_serial is None else max_serial + 1

                self.conn.execute(
                    "INSERT INTO processed_records (code, serial) VALUES (?, ?)",
                    (code, serial),
                )
                results.append((code, serial))

            self.commit_transaction()
            return results

        except Exception as e:
            self.rollback_transaction()
            raise e

    def close(self):
        """关闭数据库连接"""
        if self.conn:
            self.conn.close()
            self.conn = None


class CellProcessor:
    """单元格处理函数管理"""

    def __init__(self):
        self.functions = {}
        self.register_default_functions()

    def register_function(self, index: int, func: Callable[[Cell, Worksheet], str]):
        """注册处理函数"""
        self.functions[index] = func

    def get_function(self, index: int) -> Callable[[Cell, Worksheet], str]:
        """获取处理函数"""
        return self.functions.get(index)

    def register_default_functions(self):
        """注册默认处理函数"""
        # 函数0: 原样返回单元格值
        self.register_function(
            0, lambda cell, sheet: str(cell.value) if cell.value is not None else ""
        )

        # 函数1: 日期格式化处理
        def date_processor(cell: Cell, sheet: Worksheet) -> str:
            value = cell.value
            if isinstance(value, datetime):
                return value.strftime("%y%m%d")  # 改为 %y 获取两位年份
            try:
                # 尝试解析Excel中的日期格式
                if isinstance(value, str):
                    for fmt in ("%Y/%m/%d", "%Y-%m/%d", "%m/%d/%Y", "%m-%d-%Y"):
                        try:
                            dt = datetime.strptime(value, fmt)
                            return dt.strftime("%y%m%d")  # 改为 %y 获取两位年份
                        except ValueError:
                            continue
                # 尝试处理Excel日期序列号
                elif isinstance(value, (int, float)):
                    try:
                        dt = datetime.fromordinal(
                            datetime(1900, 1, 1).toordinal() + int(value) - 2
                        )
                        return dt.strftime("%y%m%d")  # 改为 %y 获取两位年份
                    except (ValueError, OverflowError):
                        pass
            except Exception:
                pass
            return str(value)

        self.register_function(1, date_processor)


class RowResultBuilder:
    """行结果拼接器"""

    def __init__(self, separators: List[str]):
        self.separators = separators

    def build(self, row_results: List[str]) -> str:
        """构建最终的行结果字符串"""
        if not row_results:
            return ""

        # 构建结果时处理分隔符列表长度不足的情况
        parts = [row_results[0]]
        for i in range(1, len(row_results)):
            if i <= len(self.separators):
                parts.append(self.separators[i - 1])
            parts.append(row_results[i])

        return "".join(parts)


class RowProcessor:
    """行处理器"""

    def __init__(
        self,
        cell_processor: CellProcessor,
        func_indices: List[int],
        separator_builder: RowResultBuilder,
        db_manager: DatabaseManager,
    ):
        self.cell_processor = cell_processor
        self.func_indices = func_indices
        self.separator_builder = separator_builder
        self.db_manager = db_manager
        self.results = []
        self.codes = []

    def process_row(self, row: List[Cell], sheet: Worksheet) -> str:
        """处理单行并返回编码"""
        row_results = []

        # 处理每个单元格
        for idx, cell in enumerate(row[: len(self.func_indices)]):
            func_index = self.func_indices[idx]
            processor = self.cell_processor.get_function(func_index)
            if processor:
                try:
                    row_results.append(processor(cell, sheet))
                except Exception as e:
                    row_results.append(f"!ERROR({str(e)})")
            else:
                row_results.append(str(cell.value) if cell.value is not None else "")

        # 构建最终编码
        final_code = self.separator_builder.build(row_results)
        self.codes.append(final_code)
        return final_code

    def commit_records(
        self, begin_serial: Optional[int] = None
    ) -> List[Tuple[str, int]]:
        """提交所有记录到数据库"""
        try:
            records = self.db_manager.insert_records_with_begin_serial(
                self.codes, begin_serial
            )
            self.results.extend(records)
            return records
        except Exception as e:
            self.codes = []  # 清空缓存
            raise e


class ExcelProcessor:
    """Excel处理器"""

    def __init__(
        self,
        file_path: str,
        func_indices: List[int],
        separators: List[str],
        db_path: Optional[str] = None,
        num_zill=3,
    ):
        self.file_path = file_path
        self.func_indices = func_indices
        self.separators = separators
        self.db_manager = DatabaseManager(db_path) if db_path else DatabaseManager()
        self.cell_processor = CellProcessor()
        self.row_processor = RowProcessor(
            cell_processor=self.cell_processor,
            func_indices=func_indices,
            separator_builder=RowResultBuilder(separators),
            db_manager=self.db_manager,
        )
        self.processed_records = []
        self.num_zill = num_zill

    def process_range(
        self,
        start_cell: str,
        end_cell: str,
        begin_serial: Optional[int] = None,
        sheet_name: Optional[str] = None,
    ):
        """处理指定范围的数据

        Args:
            start_cell: 起始单元格
            end_cell: 结束单元格
            begin_serial: 指定的起始流水号
            sheet_name: 要处理的工作表名称，如果为None则使用活动工作表

        Returns:
            处理后的记录列表
        """
        wb = load_workbook(self.file_path)

        # 获取指定工作表或活动工作表
        if sheet_name is not None:
            sheet = wb[sheet_name]
        else:
            sheet = wb.active

        # 获取处理范围
        start_col, start_row = self._parse_cell_ref(start_cell)
        end_col, end_row = self._parse_cell_ref(end_cell)

        # 遍历每行并处理
        for row_idx in range(start_row, end_row + 1):
            row_cells = []
            for col_idx in range(start_col, end_col + 1):
                row_cells.append(sheet.cell(row=row_idx, column=col_idx))

            # 处理行
            try:
                self.row_processor.process_row(row_cells, sheet)
            except Exception as e:
                wb.close()
                raise e

        # 提交所有记录
        try:
            records = self.row_processor.commit_records(begin_serial)
            self.processed_records.extend(records)
        except Exception as e:
            wb.close()
            raise e

        wb.close()
        return self.processed_records

    def _parse_cell_ref(self, cell_ref: str) -> Tuple[int, int]:
        """将单元格引用转换为行列索引"""
        col_str = ""
        row_str = ""
        for char in cell_ref:
            if char.isalpha():
                col_str += char
            elif char.isdigit():
                row_str += char

        # 转换列字母为数字（A=1, B=2, ...）
        col_num = 0
        for char in col_str:
            col_num = col_num * 26 + (ord(char.upper()) - ord("A")) + 1

        return (col_num, int(row_str))

    def close(self):
        """关闭资源"""
        if hasattr(self, "db_manager"):
            self.db_manager.close()

    def pad_number(self, num: int, length: int):
        """数字补零"""
        return str(int(num)).zfill(length)

    def format_results(self, records: any) -> str:
        """格式化输出结果
        Args:
            records: 需要格式化的记录列表，格式为 [(code, serial), ...]
        Returns:
            返回用换行符分隔的字符串，每行格式为 code + 补零后的serial
        """
        result_lines = []
        for code, serial in records:
            # 将每条记录格式化为 code + 补零的serial，并添加到结果列表
            result_lines.append(f"{code}{self.pad_number(serial, self.num_zill)}")

        # 用换行符连接所有行并返回
        return "\n".join(result_lines)


def split_excel_range_str(data: str) -> List[str]:
    """分割Excel范围字符串"""
    return data.split(":")


class ToolConfig(BaseModel):
    """工具配置模型，使用Pydantic进行数据验证"""

    excel_file: str
    range: str = Field(
        ..., pattern=r"^[A-Za-z]+\d+:[A-Za-z]+\d+$"
    )  # 使用正则表达式验证格式
    function_indices: List[int] = Field(..., min_items=1)
    separators: List[str]
    database_path: str
    begin_serial: Optional[int] = Field(..., ge=0)  # 大于等于0
    sheet_name: Optional[str] = "生成单号"
    num_zill: int = Field(..., ge=0)  # 输出格式化时流水号补位数

    @field_validator("separators")
    @classmethod
    def validate_separators_length(
        cls, v: List[str], info: FieldValidationInfo
    ) -> List[str]:
        """验证分隔符数量是否与函数索引数量匹配"""
        if "function_indices" in info.data and len(v) != len(
            info.data["function_indices"]
        ):
            raise ValueError("分隔符数量必须与函数索引数量相同")
        return v

def run_tool(config: ToolConfig):
    """运行工具的主函数"""
    p = ExcelProcessor(
        file_path=config.excel_file,
        func_indices=config.function_indices,
        separators=config.separators,
        db_path=config.database_path,
        num_zill=config.num_zill,
    )
    records = p.process_range(
        start_cell=split_excel_range_str(config.range)[0],
        end_cell=split_excel_range_str(config.range)[1],
        begin_serial=config.begin_serial,
        sheet_name=config.sheet_name,
    )
    return p.format_results(records)


# 使用示例
if __name__ == "__main__":
    # 配置参数
    config = ToolConfig(
        excel_file=r"d:\鼎捷项目\_纳芯微\数据导入\apmt111_icd 已入库未收票的委外工单 -麦歌恩电子NS22-补充长料号 -Ver1.1\0_单头_wrok.xlsx",
        range="A1:B880",
        function_indices=[0, 1],
        separators=["", ""],
        database_path="src/test/encoder/processing_records.db",
        begin_serial=100,  # 指定起始流水号为100
        sheet_name="单号生成",
        num_zill=4,
    )
    print(run_tool(config))
