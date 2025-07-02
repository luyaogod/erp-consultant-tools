import openpyxl
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import column_index_from_string
import os
import json


class MatchHandler:
    """
    匹配处理器基类，定义匹配/未匹配时的处理接口

    子类应实现以下方法：
    - on_match: 匹配成功时的处理逻辑
    - on_no_match: 匹配失败时的处理逻辑
    """

    def on_match(self, cell: Cell, lookup_value) -> None:
        """
        匹配成功时的默认处理：用查找值替换单元格值

        :param cell: 目标单元格对象
        :param lookup_value: 查找表中匹配的值
        """
        cell.value = lookup_value

    def on_no_match(self, cell: Cell) -> None:
        """
        匹配失败时的默认处理：不执行任何操作

        :param cell: 目标单元格对象
        """
        pass


class EmptyOrKeep(MatchHandler):
    """
    未找到匹配时的处理策略实现类

    提供两种处理方式：
    - empty: 将未匹配的单元格置空（默认）
    - keep: 保留原值不变
    """

    def __init__(self, on_no_match_action="empty"):
        """
        初始化处理策略

        :param on_no_match_action: 未匹配时的处理方式，"empty"或"keep"
        """
        self.not_found_action = on_no_match_action

    def on_no_match(self, cell: Cell) -> None:
        """
        根据策略处理未匹配单元格

        :param cell: 目标单元格对象
        """
        if self.not_found_action == "empty":
            cell.value = None  # 置空处理
        # 如果是"keep"则不做任何操作（保留原值）


class Smap:
    """
    Excel跨文件VLOOKUP处理器

    核心功能：
    1. 根据配置文件加载查找表数据
    2. 在目标文件中执行列替换
    3. 支持自定义匹配/未匹配处理策略

    设计特点：
    - 采用策略模式分离匹配处理逻辑
    - 支持多Sheet处理
    - 自动跳过指定行数
    - 可扩展的保存命名规则
    """

    def __init__(
        self,
        target_path: str,
        lookup_path: str,
        header_row: int,
        skip_rows: int = 0,
        config: dict | str = None,
        sheet_names: list = None,
        suffix: str = "_processed",
        match_handler: MatchHandler = None,
    ):
        """
        初始化处理器

        :param target_path: 待处理Excel文件路径
        :param lookup_path: 查找表Excel文件路径
        :param header_row: 列头所在行号（1-based）
        :param skip_rows: 跳过行数（从列头行之后开始计算）
        :param config: 定义字段和查找范围（优先级高于JSON配置）
        :param sheet_names: 指定处理的Sheet名称列表，None表示处理所有Sheet
        :param suffix: 输出文件后缀（默认添加'_processed'）
        :param match_handler: 自定义匹配处理器实例，None则使用默认处理器
        """
        # 初始化参数
        self.target_path = target_path
        self.lookup_path = lookup_path
        if not config:
            raise ValueError("请输字段和查找范围配置")
        if type(config) is str:
            config: dict = json.loads(config)
        self.config: dict = config
        self.header_row = header_row
        self.skip_rows = skip_rows
        self.sheet_names = sheet_names if sheet_names else []  # 空列表表示处理所有Sheet
        self.suffix = suffix

        # 初始化处理程序
        self.match_handler = match_handler or EmptyOrKeep()  # 默认使用EmptyOrKeep策略

        # 运行时数据
        self.lookup_data = {}  # 存储加载的查找表数据
        self.target_wb = None  # 目标工作簿对象

    def process(self) -> str:
        """
        执行完整处理流程

        :return: 处理后的文件保存路径
        """
        self._load_lookup_data()
        self._process_target_file()
        return self._save_processed_file()

    def _load_lookup_data(self) -> None:
        """
        加载查找表数据到内存

        根据config中的配置，从查找表中提取数据并存储为字典格式
        """
        # 以只读模式加载查找表（提高大文件读取性能）
        lookup_wb = openpyxl.load_workbook(
            self.lookup_path, read_only=True, data_only=True
        )

        for field, range_str in self.config.items():
            # 解析范围字符串（如"A2:B2853"）
            start_end = range_str.split(":")
            start_col = column_index_from_string(start_end[0][0])  # 起始列字母转数字
            start_row = int(start_end[0][1:])  # 起始行号
            end_col = column_index_from_string(start_end[1][0])  # 结束列字母转数字
            end_row = int(start_end[1][1:])  # 结束行号

            # 读取数据并保持首次出现的键值（去重）
            lookup_sheet = lookup_wb.active  # 默认使用活动工作表
            field_dict = {}
            for row in lookup_sheet.iter_rows(
                min_row=start_row,
                max_row=end_row,
                min_col=start_col,
                max_col=end_col,
                values_only=True,  # 只获取值，不保留单元格对象
            ):
                key, value = row[0], row[1]  # 假设每行两列：键和值
                if key not in field_dict:  # 只保留第一次出现的键值
                    field_dict[key] = value

            self.lookup_data[field] = field_dict  # 存储字段对应的查找字典
        lookup_wb.close()  # 关闭查找表工作簿

    def _process_target_file(self) -> None:
        """
        处理目标文件

        根据配置替换目标文件中的指定列
        """
        # 加载目标工作簿
        self.target_wb = openpyxl.load_workbook(self.target_path)
        # 确定要处理的Sheet列表
        sheets = self.sheet_names if self.sheet_names else self.target_wb.sheetnames

        for sheet_name in sheets:
            if sheet_name not in self.target_wb.sheetnames:
                continue  # 跳过不存在的Sheet

            sheet = self.target_wb[sheet_name]
            for field in self.config.keys():
                # 查找目标列位置
                target_col = self._find_target_column(sheet, field)
                if target_col is None:
                    continue  # 跳过不存在的字段

                # 处理该列数据
                self._process_column(sheet, field, target_col)

    def _find_target_column(self, sheet: Worksheet, field: str) -> int:
        """
        在目标工作表中查找指定字段的列号

        :param sheet: 目标工作表对象
        :param field: 要查找的字段名
        :return: 列号（1-based），未找到返回None
        """
        for cell in sheet[self.header_row]:  # 在表头行中查找
            if cell.value == field:
                return cell.column
        return None

    def _process_column(self, sheet: Worksheet, field: str, target_col: int) -> None:
        """
        处理单个列的数据替换

        :param sheet: 目标工作表对象
        :param field: 当前处理的字段名
        :param target_col: 目标列号（1-based）
        """
        # 计算数据起始行（表头行 + 跳过的行数 + 1）
        start_row = self.header_row + self.skip_rows + 1
        # 获取该字段对应的查找字典
        lookup_map: dict = self.lookup_data.get(field, {})

        # 遍历目标列的所有单元格
        for row in sheet.iter_rows(
            min_row=start_row, min_col=target_col, max_col=target_col
        ):
            cell = row[0]
            # 查找匹配值
            lookup_value = lookup_map.get(cell.value)
            if lookup_value is not None:
                self.match_handler.on_match(cell, lookup_value)  # 匹配成功处理
            else:
                self.match_handler.on_no_match(cell)  # 匹配失败处理

    def _save_processed_file(self) -> str:
        """
        保存处理后的文件

        :return: 新文件的保存路径
        """
        # 构造新文件名（原文件名 + 后缀）
        original_dir = os.path.dirname(self.target_path)
        original_name = os.path.basename(self.target_path)
        base_name, ext = os.path.splitext(original_name)
        new_path = os.path.join(original_dir, f"{base_name}{self.suffix}{ext}")

        # 保存并关闭工作簿
        self.target_wb.save(new_path)
        self.target_wb.close()
        return new_path


def smap(
    target_path: str,
    lookup_path: str,
    config: dict = None,
    header_row: int = 2,
    skip_rows: int = 0,
    sheet_names: list = None,
    suffix: str = "_processed",
    not_math: str = "empty",
) -> str:
    """
    快捷函数：创建Smap实例并执行处理

    :param target_path: 目标文件路径
    :param lookup_path: 查找表文件路径
    :param config: 字段映射配置
    :param header_row: 表头行号
    :param skip_rows: 跳过的行数
    :param sheet_names: 指定处理的Sheet列表
    :param suffix: 输出文件后缀
    :param not_math: 未匹配时的处理方式（"empty"或"keep"）
    :return: 处理后的文件路径
    """
    smap = Smap(
        target_path=target_path,
        lookup_path=lookup_path,
        config=config,
        header_row=header_row,
        skip_rows=skip_rows,
        sheet_names=sheet_names,
        suffix=suffix,
        match_handler=EmptyOrKeep(not_math),
    )
    return smap.process()


# 使用示例
if __name__ == "__main__":
    # 示例调用
    result_path = smap(
        target_path="src/test/smap/维护BOM信息all250427.xlsx",
        lookup_path="src/test/smap/昆山1对照表_no_duplicates.xlsx",
        header_row=2,
        config='{"元件品号": "A2:B2853"}',
    )

    print(f"处理完成，文件已保存至：{result_path}")
