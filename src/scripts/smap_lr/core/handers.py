import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell import Cell


class CopyCloumnHandler:
    def __init__(self, max_num: int):
        self.max_num = max_num
        self.matches = []
        self.not_matches = []

    def my_on_match(
        self, cell1: Cell, cell2: Cell, sheet1: Worksheet, sheet2: Worksheet
    ) -> None:
        """
        :param cell1: 遍历表cell (Sheet3)
        :param cell2: 查询表cell (Sheet2)
        :param sheet1: 遍历表 (Sheet3)
        :param sheet2: 查询表 (Sheet2)
        """
        self.matches.append(f"{cell1.value}")
        # 将匹配的单元格标为绿色
        cell1.fill = openpyxl.styles.PatternFill(
            start_color="00FF00", fill_type="solid"
        )
        cell2.fill = openpyxl.styles.PatternFill(
            start_color="00FF00", fill_type="solid"
        )

        # 获取参照表cell所在的列
        ref_col = cell2.column

        # 获取遍历表cell所在的列
        target_col = cell1.column

        # 获取参照表的数据范围（从匹配行+1开始，共max_num行）
        start_row = cell2.row + 1
        end_row = start_row + self.max_num - 1

        # 从参照表复制数据到遍历表
        for row_idx, row in enumerate(range(start_row, end_row + 1), start=1):
            # 获取参照表单元格的值
            ref_cell = sheet2.cell(row=row, column=ref_col)

            # 设置遍历表对应单元格的值（从匹配行+1开始）
            target_row = cell1.row + row_idx
            if target_row <= sheet1.max_row:
                if not ref_cell.value:
                    sheet1.cell(row=target_row, column=target_col).value = None
                else:
                    sheet1.cell(row=target_row, column=target_col, value=ref_cell.value)
            else:
                # 如果超出当前行数，则追加新行
                sheet1.cell(row=target_row, column=target_col, value=ref_cell.value)

    def my_on_nomatch(self, cell: Cell, sheet1: Worksheet, sheet2: Worksheet) -> None:
        """
        :param cell: 遍历表cell
        :param sheet1: 遍历表
        :param sheet2: 查询表
        """
        self.not_matches.append(f"{cell.value}")
        # 将不匹配的单元格标为红色
        cell.fill = openpyxl.styles.PatternFill(start_color="FF0000", fill_type="solid")
