import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell import Cell
from openpyxl.utils.exceptions import InvalidFileException
from typing import Optional, Dict, List, Callable, Tuple, Any


class Processor:
    def __init__(
        self,
        file_path: str,
        sheet1_name: str = "Sheet1",
        sheet2_name: str = "Sheet2",
        sheet3_name: str = "Sheet3",
    ) -> None:
        """
        初始化Excel处理器
        :param file_path: Excel文件路径
        :param sheet1_name: 第一个工作表名称（默认Sheet1）
        :param sheet2_name: 第二个工作表名称（默认Sheet2）
        :param sheet3_name: 第三个工作表名称（默认Sheet3）
        """
        self.file_path: str = file_path
        self.sheet_names: Dict[str, str] = {
            "sheet1": sheet1_name,
            "sheet2": sheet2_name,
            "sheet3": sheet3_name,
        }
        self.wb: Optional[Workbook] = None
        self.sheet1: Optional[Worksheet] = None
        self.sheet2: Optional[Worksheet] = None
        self.sheet3: Optional[Worksheet] = None

        # 加载工作簿和工作表
        self._load_workbook()
        self._prepare_sheets()

    def _load_workbook(self) -> None:
        """加载Excel工作簿"""
        try:
            self.wb = openpyxl.load_workbook(self.file_path)
        except (FileNotFoundError, InvalidFileException) as e:
            raise Exception(f"文件加载失败: {str(e)}")
        except Exception as e:
            raise Exception(f"未知错误: {str(e)}")

    def _prepare_sheets(self) -> None:
        """准备所需的工作表"""
        if self.wb is None:
            raise RuntimeError("工作簿未初始化")

        existing_sheets: List[str] = self.wb.sheetnames

        # 检查必须存在的工作表
        if self.sheet_names["sheet1"] not in existing_sheets:
            raise ValueError(f"工作表 '{self.sheet_names['sheet1']}' 不存在")
        if self.sheet_names["sheet2"] not in existing_sheets:
            raise ValueError(f"工作表 '{self.sheet_names['sheet2']}' 不存在")

        # 创建或获取sheet3
        if self.sheet_names["sheet3"] in existing_sheets:
            self.sheet3 = self.wb[self.sheet_names["sheet3"]]
        else:
            self.sheet3 = self.wb.create_sheet(title=self.sheet_names["sheet3"])

        # 设置工作表对象
        self.sheet1 = self.wb[self.sheet_names["sheet1"]]
        self.sheet2 = self.wb[self.sheet_names["sheet2"]]

        # 复制sheet1到sheet3
        self._copy_sheet(self.sheet1, self.sheet3)

    def _copy_sheet(self, source_sheet: Worksheet, target_sheet: Worksheet) -> None:
        """
        复制工作表内容（带格式）
        :param source_sheet: 源工作表
        :param target_sheet: 目标工作表
        """
        # 清空目标工作表（如果已有内容）
        if target_sheet.dimensions:
            target_sheet.delete_rows(1, target_sheet.max_row)
            target_sheet.delete_cols(1, target_sheet.max_column)

        # 复制所有单元格（带值）
        for row in source_sheet.iter_rows():
            row_data: List[Any] = [cell.value for cell in row]
            target_sheet.append(row_data)

        # 复制列宽
        for col_letter, col_dim in source_sheet.column_dimensions.items():
            if col_dim.width is not None:
                target_sheet.column_dimensions[col_letter].width = col_dim.width

    def left_to_right(
        self,
        row1: int,
        row2: int,
        on_match: Callable[[Cell, Cell, Worksheet, Worksheet], None],
        on_nomatch: Callable[[Cell, Worksheet], None],
    ) -> None:
        """
        遍历方法1：从Sheet3的row1行查找Sheet2的row2行
        :param row1: Sheet3中的行号
        :param row2: Sheet2中的行号
        :param on_match: 匹配时的回调函数(cell_sheet3, cell_sheet2)
        :param on_nomatch: 不匹配时的回调函数(cell_sheet3)
        """
        if self.sheet2 is None or self.sheet3 is None:
            raise RuntimeError("工作表未初始化")

        # 获取Sheet2目标行
        sheet2_row: List[Tuple[Cell, ...]] = list(
            self.sheet2.iter_rows(min_row=row2, max_row=row2, values_only=False)
        )
        if not sheet2_row:
            raise ValueError(f"Sheet2中行{row2}不存在")
        sheet2_row_cells: Tuple[Cell, ...] = sheet2_row[0]
        sheet2_values: List[Any] = [cell.value for cell in sheet2_row_cells]

        # 获取Sheet3目标行
        sheet3_row: List[Tuple[Cell, ...]] = list(
            self.sheet3.iter_rows(min_row=row1, max_row=row1, values_only=False)
        )
        if not sheet3_row:
            raise ValueError(f"Sheet3中行{row1}不存在")

        # 遍历Sheet3的目标行
        for cell in sheet3_row[0]:
            try:
                # 查找匹配值在Sheet2中的位置
                match_index: int = sheet2_values.index(cell.value)
                on_match(cell, sheet2_row_cells[match_index], self.sheet3, self.sheet2)
            except ValueError:
                on_nomatch(cell, self.sheet3, self.sheet2)

    def right_to_left(
        self,
        row2: int,
        row1: int,
        on_match: Callable[[Cell, Cell, Worksheet, Worksheet], None],
        on_nomatch: Callable[[Cell, Worksheet, Worksheet], None],
    ) -> None:
        """
        遍历方法2：从Sheet2的row2行查找Sheet3的row1行
        :param row2: Sheet2中的行号
        :param row1: Sheet3中的行号
        :param on_match: 匹配时的回调函数(cell_sheet3, cell_sheet2)
        :param on_nomatch: 不匹配时的回调函数(cell_sheet2)
        """
        if self.sheet2 is None or self.sheet3 is None:
            raise RuntimeError("工作表未初始化")

        # 获取Sheet3目标行
        sheet3_row: List[Tuple[Cell, ...]] = list(
            self.sheet3.iter_rows(min_row=row1, max_row=row1, values_only=False)
        )
        if not sheet3_row:
            raise ValueError(f"Sheet3中行{row1}不存在")
        sheet3_row_cells: Tuple[Cell, ...] = sheet3_row[0]
        sheet3_values: List[Any] = [cell.value for cell in sheet3_row_cells]

        # 获取Sheet2目标行
        sheet2_row: List[Tuple[Cell, ...]] = list(
            self.sheet2.iter_rows(min_row=row2, max_row=row2, values_only=False)
        )
        if not sheet2_row:
            raise ValueError(f"Sheet2中行{row2}不存在")

        # 遍历Sheet2的目标行
        for cell in sheet2_row[0]:
            try:
                # 查找匹配值在Sheet3中的位置
                match_index: int = sheet3_values.index(cell.value)
                on_match(sheet3_row_cells[match_index], cell, self.sheet2, self.sheet3)
            except ValueError:
                on_nomatch(cell, self.sheet2, self.sheet3)

    def save(self, output_path: Optional[str] = None) -> None:
        """
        保存工作簿
        :param output_path: 输出路径（如果为None则覆盖原文件）
        """
        if self.wb is None:
            raise RuntimeError("工作簿未初始化")

        save_path = output_path if output_path else self.file_path
        self.wb.save(save_path)

    def process(
        self,
        direction: bool,
        row2: int,
        row1: int,
        on_match: Callable[[Cell, Cell, Worksheet, Worksheet], None],
        on_nomatch: Callable[[Cell, Worksheet, Worksheet], None],
    ):
        if direction:
            self.left_to_right(row2, row1, on_match, on_nomatch)
        else:
            self.right_to_left(row2, row1, on_match, on_nomatch)




    