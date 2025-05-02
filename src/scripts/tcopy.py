import os
from openpyxl import Workbook


def tcopy(input_file_path, suffix="_clone", output_path=None):
    """
    复刻Excel文件的所有Sheet（创建空副本）

    参数:
        input_file_path (str): 输入Excel文件路径
        suffix (str): 输出文件名后缀，默认为"_clone"
        output_path (str): 输出路径，默认为输入文件同路径

    返回:
        str: 生成的副本文件路径
    """
    # 加载原始工作簿仅获取sheet名
    from openpyxl import load_workbook

    original_wb = load_workbook(input_file_path, read_only=True)
    sheet_names = original_wb.sheetnames
    original_wb.close()

    # 创建新工作簿
    new_wb = Workbook()

    # 删除默认创建的Sheet（如果有）
    if "Sheet" in new_wb.sheetnames:
        del new_wb["Sheet"]

    # 创建所有同名Sheet
    for name in sheet_names:
        new_wb.create_sheet(name)

    # 确定输出路径和文件名
    if output_path is None:
        output_path = os.path.dirname(input_file_path)

    original_name = os.path.splitext(os.path.basename(input_file_path))[0]
    output_filename = f"{original_name}{suffix}.xlsx"
    output_file_path = os.path.join(output_path, output_filename)

    # 保存新工作簿
    new_wb.save(output_file_path)

    return output_file_path


if __name__ == "__main__":
# 示例用法
    cloned_file = tcopy("src/test/tcopy/维护BOM信息all250427.xlsx", suffix="_template")
    print(f"已创建副本文件: {cloned_file}")
